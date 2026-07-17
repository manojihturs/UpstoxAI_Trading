"""
trade_export.py
Local-only automatic backup of every closed trade into a single Excel
workbook, one sheet per calendar day (sheet name = "YYYY-MM-DD"). This is
purely a durable, human-readable copy on whatever disk engine.py happens to
be running on -- trading_state.db (via state_store.py) remains the actual
source of truth the dashboard reads from. On an ephemeral host this file
is just as short-lived as the database; it's meant for local runs where
the disk persists across restarts.
"""
import os
import datetime
from openpyxl import Workbook, load_workbook

from config import PATHS

EXCEL_PATH = os.path.join(os.path.dirname(PATHS["DB_PATH"]), "trade_history.xlsx")

COLUMNS = [
    "entry_time", "instrument", "direction", "strike", "qty",
    "entry_ltp_raw", "entry_ltp_net", "exit_time", "exit_ltp_raw", "exit_ltp_net",
    "exit_reason", "gross_pnl", "costs_total", "net_pnl",
]


def _sheet_name_for(exit_time_iso):
    """'2026-07-14T15:20:00' -> '2026-07-14' (valid Excel sheet name)."""
    return exit_time_iso.split("T")[0]


def append_closed_trade(position, path=EXCEL_PATH):
    """position: a closed position dict (from state_store), with all of
    COLUMNS populated. Appends one row to that trade's exit-day sheet,
    creating the workbook and/or sheet as needed."""
    sheet_name = _sheet_name_for(position["exit_time"])

    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # drop the default blank sheet

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(COLUMNS)
    else:
        ws = wb[sheet_name]

    ws.append([position.get(col) for col in COLUMNS])
    wb.save(path)
