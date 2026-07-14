"""
test_trade_export.py
Tests the daily-sheet Excel backup against a throwaway workbook path (never
the real trade_history.xlsx). Run with: pytest test_trade_export.py
"""
import openpyxl
import pytest

import trade_export


def make_trade(exit_time, net_pnl=100.0):
    return {
        "entry_time": "2026-07-14T10:00:00",
        "instrument": "NIFTY",
        "direction": "CE",
        "strike": 24800,
        "qty": 75,
        "entry_ltp_raw": 120.0,
        "entry_ltp_net": 120.6,
        "exit_time": exit_time,
        "exit_ltp_raw": 130.0,
        "exit_ltp_net": 129.4,
        "exit_reason": "TARGET",
        "gross_pnl": 660.0,
        "costs_total": 55.0,
        "net_pnl": net_pnl,
    }


def test_creates_workbook_and_sheet_for_new_trade(tmp_path):
    path = str(tmp_path / "trade_history.xlsx")
    trade_export.append_closed_trade(make_trade("2026-07-14T15:20:00"), path=path)

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["2026-07-14"]
    ws = wb["2026-07-14"]
    assert ws.cell(row=1, column=1).value == "entry_time"  # header row
    assert ws.cell(row=2, column=2).value == "NIFTY"  # instrument column


def test_appends_multiple_trades_same_day_to_same_sheet(tmp_path):
    path = str(tmp_path / "trade_history.xlsx")
    trade_export.append_closed_trade(make_trade("2026-07-14T11:00:00", net_pnl=50.0), path=path)
    trade_export.append_closed_trade(make_trade("2026-07-14T14:00:00", net_pnl=-30.0), path=path)

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["2026-07-14"]
    ws = wb["2026-07-14"]
    assert ws.max_row == 3  # header + 2 trades


def test_different_days_get_separate_sheets(tmp_path):
    path = str(tmp_path / "trade_history.xlsx")
    trade_export.append_closed_trade(make_trade("2026-07-14T15:00:00"), path=path)
    trade_export.append_closed_trade(make_trade("2026-07-15T10:30:00"), path=path)

    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == {"2026-07-14", "2026-07-15"}
    assert wb["2026-07-14"].max_row == 2  # header + 1 trade
    assert wb["2026-07-15"].max_row == 2


if __name__ == "__main__":
    import sys
    sys.exit(pytest.main([__file__, "-v"]))
